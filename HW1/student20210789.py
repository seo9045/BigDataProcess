#!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook('student.xlsx')
ws = wb.active

ws.cell(row=1, column=7).value = 'Total'
ws.cell(row=1, column=8).value = 'Grade'

total_list = []

for row in range(2, ws.max_row + 1):
    midterm = ws.cell(row=row, column=3).value
    final = ws.cell(row=row, column=4).value
    homework = ws.cell(row=row, column=5).value
    attendance = ws.cell(row=row, column=6).value

    total = (midterm * 0.3) + (final * 0.35) + (homework * 0.34) + (attendance * 0.01)
    ws.cell(row=row, column=7).value = total
    total_list.append(total)

total_list.sort(reverse=True)

for row in range(2, ws.max_row + 1):
    total = ws.cell(row=row, column=7).value

    if total < 40:
        grade = 'F'
    else:
        score_a = int(len(total_list) * 0.3)
        score_b = int(len(total_list) * 0.7)
        score_a_plus = min(int(score_a * 0.5), len(total_list) - score_a)
        score_b_plus = min(int(score_b * 0.5), len(total_list) - score_a - score_b)

        if total in total_list[:score_a]:
            grade = 'A'
        elif total in total_list[score_a:score_a + score_a_plus]:
            grade = 'A+'
        elif total in total_list[score_a + score_a_plus:score_a + score_b]:
            grade = 'B'
        elif total in total_list[score_a + score_b:score_a + score_b + score_b_plus]:
            grade = 'B+'
        elif total in total_list[score_a + score_b + score_b_plus:]:
            grade = 'C0'
        else:
            grade = 'C+'

    ws.cell(row=row, column=8).value = grade

wb.save("student_total.xlsx")
wb.close()
